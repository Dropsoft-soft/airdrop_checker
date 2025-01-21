
import os
import json, time
from loguru import logger
from openpyxl import Workbook
from datetime import datetime
import openpyxl
import asyncio
from loguru import logger
import time
import json
import aiohttp
from scr.helper import WALLETS, ADRESSESS

TIMEOUT = [10, 40]
MAX_RETRY = 4
ERROR_CODE_EXCEPTION = -1
ERROR_CODE_FAILED_REQUEST = -2


class Request_main():
    def __init__(self, file_name: str, use_addresses: bool = True):
        self.file_name = file_name
        self.success_array = {}
        self.use_addresses = use_addresses

    async def global_request(self, method="get", request_retry=0, need_sleep= False, **kwargs):
        async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=False)) as session:
            if request_retry > MAX_RETRY:
                return
            retry = 0

            while True:
                try:
                    if method == "post":
                        response = await session.post(**kwargs)
                    elif method == "get":
                        response = await session.get(**kwargs)
                    elif method == "put":
                        response = await session.put(**kwargs)
                    elif method == "options":
                        response = await session.options(**kwargs)

                    status_code = response.status
                    if status_code == 201 or status_code == 200:
                        try:
                            return status_code, await response.json()
                        except json.decoder.JSONDecodeError:
                            logger.info('The request success but not contain a JSON')
                            break
                    else:
                        retry += 1
                        if retry > 4:
                            message = f'[{kwargs["url"]}] request attempts reached max retries count'
                            logger.error(message)
                            return ERROR_CODE_FAILED_REQUEST, message

                except ConnectionError as error:
                    logger.error(f'HTTPSConnectionPool - {kwargs["url"]} failed to make request | {error}')
                    if need_sleep:
                        time.sleep(25)
                    await self.global_request(method=method, request_retry=request_retry + 1, need_sleep=True, **kwargs)
                    break

                except Exception as error:
                    logger.error(f'{kwargs["url"]} failed to make request | {error}')
                    if need_sleep:
                        time.sleep(10)
                    return ERROR_CODE_EXCEPTION, error
        

    def set_column_widths(self, sheet):
        column_widths = {
            'A': 20,  # Date Time
            'B': 50,  # Wallet
            'C': 15,  # Amount
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

    

    def add_data(self, wallet: str, amount: str, amount_name: str = 'Amount'):
        if not os.path.exists(self.file_name):
            book = Workbook()
            sheet = book.active
            sheet['A1'] = 'Date Time'
            sheet['B1'] = 'Wallet'
            sheet['C1'] = amount_name
        
            self.set_column_widths(sheet)
            book.save(self.file_name)
            book.close()

        book = openpyxl.load_workbook(self.file_name)
        sheet = book.active

        new_row = sheet.max_row + 1
        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=new_row, column=1).value = current_datetime
        sheet.cell(row=new_row, column=2).value = wallet
        sheet.cell(row=new_row, column=3).value = amount

        book.save(self.file_name)
        book.close()
        logger.success(f'Saved {wallet} amount:{amount}')

    async def run_module(self, key):
       pass

    def _async_run_module(self, key):
        asyncio.run(self.run_module(key))

    def get_wallets(self):
        w = []
        if self.use_addresses:
            w = ADRESSESS
        else:
            w = WALLETS
        wallets = [
            {
                "id": _id,
                "key": key,
            } for _id, key in enumerate(w, start=1)
        ]
        return wallets