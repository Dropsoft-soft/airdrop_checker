import os
import json, time
from loguru import logger
from openpyxl import Workbook
from datetime import datetime
import openpyxl
import asyncio
from random import uniform
from concurrent.futures import ThreadPoolExecutor
from loguru import logger
import time
import json
import aiohttp
import sys

TIMEOUT = [3, 10]
MAX_RETRY = 4
ERROR_CODE_EXCEPTION = -1
ERROR_CODE_FAILED_REQUEST = -2

'''Settings use or not use proxies and filename'''
EXEL_FILENAME = 'odos_elligable.xlsx'
USE_PROXY = True

with open("addresses.txt", "r") as f:
    WALLETS = [row.strip() for row in f]

with open("proxies.txt", "r") as f:
    PROXIES = [row.strip() for row in f]

def decimalToInt(qty, decimal):
    return qty/ int("".join((["1"]+ ["0"]*decimal)))

def get_wallet_proxies(wallets, proxies):
    try:
        result = {}
        for i in range(len(wallets)):
            result[wallets[i]] = proxies[i % len(proxies)]
        return result
    except: None

WALLET_PROXIES  = get_wallet_proxies(WALLETS, PROXIES)

async def global_request(method="get", request_retry=0, need_sleep= False, **kwargs):
    async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=False)) as session:
        # if proxy is not None:
        #     prx = {
        #             "http": f"{proxy}",
        #             "https": f"{proxy}"
        #         }

        if request_retry > MAX_RETRY:
            return
        retry = 0

        while True:
            try:
                if method == "post":
                    response = await session.post(**kwargs)
                elif method == "get":
                    response = await session.get(**kwargs)
                elif method == "put":
                    response = await session.put(**kwargs)
                elif method == "options":
                    response = await session.options(**kwargs)

                status_code = response.status
                if status_code == 201 or status_code == 200:
                    try:
                        return status_code, await response.json()
                    except json.decoder.JSONDecodeError:
                        logger.info('The request success but not contain a JSON')
                        break
                else:
                   
                    retry += 1
                    if retry > 4:
                        message = f'[{kwargs["url"]}] request attempts reached max retries count'
                        logger.error(message)
                        return ERROR_CODE_FAILED_REQUEST, message

            except ConnectionError as error:
                logger.error(f'HTTPSConnectionPool - {kwargs["url"]} failed to make request | {error}')
                if need_sleep:
                    time.sleep(25)
                await global_request(method=method, request_retry=request_retry + 1, need_sleep=True, **kwargs)
                break

            except Exception as error:
                logger.error(f'{kwargs["url"]} failed to make request | {error}')
                if need_sleep:
                    time.sleep(10)
                return ERROR_CODE_EXCEPTION, error


def set_column_widths(sheet):
    column_widths = {
        'A': 20,  # Date Time
        'B': 50,  # Wallet
        'C': 15,  # Amount
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

def add_data(wallet: str, amount: str):
    file_name = EXEL_FILENAME
    if not os.path.exists(file_name):
        book = Workbook()
        sheet = book.active
        sheet['A1'] = 'Date Time'
        sheet['B1'] = 'Wallet'
        sheet['C1'] = 'Amount'
      
        set_column_widths(sheet)
        book.save(file_name)
        book.close()

    book = openpyxl.load_workbook(file_name)
    sheet = book.active

    new_row = sheet.max_row + 1
    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet.cell(row=new_row, column=1).value = current_datetime
    sheet.cell(row=new_row, column=2).value = wallet
    sheet.cell(row=new_row, column=3).value = amount

    book.save(file_name)
    book.close()
    logger.success(f'Saved {wallet} amount:{amount}')

# def start_check_marks(wallet):
#     proxy = None
#     if USE_PROXY:
#         proxy = WALLET_PROXIES[wallet]
#     lowercased = wallet.lower()
#     url = f'https://trailblazer.hekla.taiko.xyz/api/address?address={lowercased}'

#     # response = request(url=url, wallet=wallet, proxy=proxy)
#     # if response['value'] != None:
#     #     add_data(wallet, response['value'])

success_array = {}

async def run_module(key):
    try:
        proxy = None
        if USE_PROXY:
            proxy = WALLET_PROXIES[key]
        lowercased = key.lower()
        url = f'https://api.odos.xyz/loyalty/users/{lowercased}/balances'
        headers = {
            'accept': '*/*',
            'accept-language': 'en-US,en;q=0.9,uk;q=0.8',
            'cache-control': 'no-cache',
            'origin': 'https://app.odos.xyz',
            'pragma': 'no-cache',
            'priority': 'u=1, i',
            'referer': 'https://app.odos.xyz/',
            'sec-ch-ua': '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"macOS"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36'
        }
        status, result = await global_request(url=url, headers=headers, proxy=proxy)
        if result['data']['pendingTokenBalance'] != None:
            val = int(result['data']['pendingTokenBalance'])
            if val > 0:
                val = decimalToInt(val, 18)
                k = f'{key}'
                logger.success(f'{k} amount:{val}')
                success_array[k] = val
            else:
                logger.warning(f'Skip {key} amount {val}')
       
    except Exception as e:
        logger.error(e)

def _async_run_module(account_id, key):
    asyncio.run(run_module(key))

def get_wallets():
    wallets = [
        {
            "id": _id,
            "key": key,
        } for _id, key in enumerate(WALLETS, start=1)
    ]
    return wallets

if __name__ == "__main__":
    if sys.platform == 'win32':
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    wallets = get_wallets()

    with ThreadPoolExecutor(max_workers=len(wallets)) as executor:
        for _, account in enumerate(wallets, start=1):
            executor.submit(
                _async_run_module,
                account.get("id"),
                account.get("key")
            )

    print(f'{len(success_array)} saving...')
    for i in success_array.keys():
        add_data(f'{i}', f'{success_array[i]}')    
    print(f'Finished')

        

